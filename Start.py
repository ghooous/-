import win32gui, win32ui, win32con
from ctypes import windll
from PIL import Image, ImageChops, ImageOps, ImageStat
import pytesseract
import time, os
from openpyxl import load_workbook

def find_window_by_title(title):
    """
    Находит HWND окна по названию (подстроке в заголовке, без учета регистра).
    Возвращает handle окна или None.
    """
    def _win_enum_callback(hwnd, result_list):
        wtitle = win32gui.GetWindowText(hwnd)
        if title.lower() in wtitle.lower():
            result_list.append(hwnd)
    results = []
    win32gui.EnumWindows(_win_enum_callback, results)
    return results[0] if results else None

def capture_window_image(hwnd, client_width, client_height):
    """
    Захватывает изображение клиентской области окна (hwnd) и возвращает PIL Image.
    Использует Win32 API для захвата содержимого вне зависимости от перекрытия окна.
    """
    hwndDC = win32gui.GetWindowDC(hwnd)
    mfcDC = win32ui.CreateDCFromHandle(hwndDC)
    saveDC = mfcDC.CreateCompatibleDC()
    saveBitMap = win32ui.CreateBitmap()
    saveBitMap.CreateCompatibleBitmap(mfcDC, client_width, client_height)
    saveDC.SelectObject(saveBitMap)
    
    try:
        # Флаг 1 = PW_CLIENTONLY (только клиентская область)
        windll.user32.PrintWindow(hwnd, saveDC.GetSafeHdc(), 1)
        bmpinfo = saveBitMap.GetInfo()
        bmpstr = saveBitMap.GetBitmapBits(True)
        # Создаем изображение PIL из сырого буфера (BGRX -> RGB)
        img = Image.frombuffer('RGB', (bmpinfo['bmWidth'], bmpinfo['bmHeight']), bmpstr, 'raw', 'BGRX', 0, 1)
    finally:
        # Освобождаем ресурсы GDI
        try:
            win32gui.ReleaseDC(hwnd, hwndDC)
            mfcDC.DeleteDC()
            saveDC.DeleteDC()
            win32gui.DeleteObject(saveBitMap.GetHandle())
        except Exception as e:
            print(f"Ошибка освобождения ресурсов GDI: {e}")
    return img

def match_image(region_img, templates):
    """
    Ищет соответствие изображения области (region_img) с шаблонами из списка templates.
    templates: список пар (name, PIL.Image) для определенной категории зон.
    Возвращает имя шаблона (без расширения), если найдено совпадение, иначе None.
    """
    region = region_img.convert('RGB')
    for name, templ in templates:
        ref_img = templ
        if templ.size != region.size:
            # Масштабируем шаблон к размеру области для сравнения
            ref_img = templ.resize(region.size, Image.ANTIALIAS).convert('RGB')
        else:
            ref_img = ref_img.convert('RGB')
        diff = ImageChops.difference(region, ref_img)
        if not diff.getbbox():
            # Идентичное изображение
            return name
        # Проверяем среднее отличие по пикселям (для учета незначительных расхождений)
        stat = ImageStat.Stat(diff)
        if isinstance(stat.mean, (list, tuple)):
            mean_diff = sum(stat.mean) / len(stat.mean)
        else:
            mean_diff = stat.mean
        if mean_diff < 5:  # порог допуска небольшого расхождения
            return name
    return None

def ocr_image(region_img, numeric=False):
    """
    Распознаёт текст на изображении region_img с помощью Tesseract OCR.
    Если numeric=True, ограничивает распознавание цифрами.
    Возвращает распознанный текст (или пустую строку, если ничего не найдено).
    """
    gray = region_img.convert('L')
    # Бинаризация изображения (порог 128)
    binary = gray.point(lambda x: 255 if x > 128 else 0, 'L')
    # Если фон тёмный (большая часть пикселей чёрные), инвертируем, чтобы получить чёрный текст на белом фоне
    stat = ImageStat.Stat(binary)
    if stat.mean[0] < 128:
        binary = ImageOps.invert(binary)
    config = "--psm 7"
    if numeric:
        config += " -c tessedit_char_whitelist=0123456789"
    text = pytesseract.image_to_string(binary, config=config)
    return text.strip()

if __name__ == "__main__":
    window_name = input("Введите название окна: ").strip()
    if not window_name:
        print("Название окна не задано.")
        exit(1)
    # Получаем HWND окна по полному или частичному совпадению заголовка
    hwnd = win32gui.FindWindow(None, window_name)
    if not hwnd:
        hwnd = find_window_by_title(window_name)
    if not hwnd:
        print(f"Окно с названием \"{window_name}\" не найдено.")
        exit(1)
    # Получаем размеры клиентской области (без рамок и заголовка)
    left, top, right, bottom = win32gui.GetClientRect(hwnd)
    client_width = right - left
    client_height = bottom - top
    if client_width == 0 or client_height == 0:
        # Если окно свернуто, попробуем восстановить
        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
        left, top, right, bottom = win32gui.GetClientRect(hwnd)
        client_width = right - left
        client_height = bottom - top
        if client_width == 0 or client_height == 0:
            print("Не удалось получить размеры клиентской области окна.")
            exit(1)
    print(f"Найдено окно \"{win32gui.GetWindowText(hwnd)}\" (HWND={hwnd}), размер клиентской области: {client_width}x{client_height}")
    # Читаем таблицу координат из Excel
    excel_path = r"C:\MLBB\data\Coordinates.xlsx"
    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"Ошибка открытия Excel-файла: {e}")
        exit(1)
    sheet_name = 'Coordinates' if 'Coordinates' in wb.sheetnames else wb.sheetnames[0]
    sheet = wb[sheet_name]
    coords_data = []
    click_rel = None
    for row in sheet.iter_rows(values_only=True):
        if not row or row[1] is None:
            continue
        # Пропускаем строку с формулой (если первая ячейка содержит код)
        if isinstance(row[0], str) and row[0].strip().startswith("def"):
            continue
        # Столбцы: [ID, Team, Group, Zone, X_Left_Rel, X_Right_Rel, Y_Top_Rel, Y_Bottom_Rel, ...]
        Team = str(row[1]) if row[1] is not None else ""
        Group = str(row[2]) if row[2] is not None else ""
        Zone = str(row[3]) if row[3] is not None else ""
        try:
            X_left_rel = float(row[4]); X_right_rel = float(row[5])
            Y_top_rel = float(row[6]); Y_bottom_rel = float(row[7])
        except (TypeError, ValueError):
            continue
        if Zone == "click_point":
            click_rel = (X_left_rel, Y_top_rel)
            continue
        coords_data.append((Team, Group, Zone, X_left_rel, X_right_rel, Y_top_rel, Y_bottom_rel))
    if not coords_data:
        print("Не удалось загрузить координаты из Excel.")
        exit(1)
    # Вычисляем средний относительный центр по X (для корректировки координат по ширине окна)
    centers = [ (x_left + x_right) / 2 for (_, _, _, x_left, x_right, _, _) in coords_data ]
    x_rel_center = sum(centers) / len(centers)
    # Формируем словарь координат в пикселях для каждой зоны
    coord_map = {}
    for (Team, Group, Zone, X_left_rel, X_right_rel, Y_top_rel, Y_bottom_rel) in coords_data:
        center_x = client_width / 2
        left_px = int(round(center_x + (X_left_rel - x_rel_center) * client_height))
        right_px = int(round(center_x + (X_right_rel - x_rel_center) * client_height))
        top_px = int(round(Y_top_rel * client_height))
        bottom_px = int(round(Y_bottom_rel * client_height))
        team_key = Team; group_key = Group if Group not in [None, "", "-"] else "_global"
        coord_map.setdefault(team_key, {}).setdefault(group_key, {})
        coord_map[team_key][group_key][Zone] = (left_px, top_px, right_px, bottom_px)
    # Координаты точки для клика (если задана в таблице)
    click_point = None
    if click_rel:
        cx, cy = click_rel
        center_x = client_width / 2
        x_click = int(round(center_x + (cx - x_rel_center) * client_height))
        y_click = int(round(cy * client_height))
        click_point = (x_click, y_click)
    # Загружаем образцы изображений для зон 'q', 'e', 's'
    data_dir = r"C:\MLBB\data"
    script_dir = os.path.join(data_dir, "data.for.script")
    ref_images = {"q": [], "e": [], "s": []}
    for zone_type in ["q", "e", "s"]:
        folder = os.path.join(script_dir, zone_type)
        if os.path.isdir(folder):
            for fname in os.listdir(folder):
                if fname.lower().endswith((".png", ".jpg", ".bmp")):
                    try:
                        img = Image.open(os.path.join(folder, fname)).convert("RGB")
                        name = os.path.splitext(fname)[0]
                        ref_images[zone_type].append((name, img))
                    except Exception as e:
                        print(f"Не удалось загрузить образец {fname}: {e}")
        else:
            print(f"Не найдена папка с образцами: {folder}")
    # Создаем выходные директории для файлов (если еще не существуют)
    for team, groups in coord_map.items():
        team_dir = os.path.join(data_dir, team)
        for group in groups:
            out_dir = team_dir if group == "_global" else os.path.join(team_dir, group)
            os.makedirs(out_dir, exist_ok=True)
    print("Начало мониторинга. Для остановки нажмите Ctrl+C.")
    try:
        while True:
            start_time = time.time()
            # Двойной клик по заданной точке (каждые 3 секунды)
            if click_point:
                x_click, y_click = click_point
                lParam = (y_click << 16) | x_click
                win32gui.SendMessage(hwnd, win32con.WM_MOUSEMOVE, 0, lParam)
                win32gui.SendMessage(hwnd, win32con.WM_LBUTTONDOWN, win32con.MK_LBUTTON, lParam)
                win32gui.SendMessage(hwnd, win32con.WM_LBUTTONUP, 0, lParam)
                time.sleep(0.3)
                win32gui.SendMessage(hwnd, win32con.WM_LBUTTONDOWN, win32con.MK_LBUTTON, lParam)
                win32gui.SendMessage(hwnd, win32con.WM_LBUTTONUP, 0, lParam)
            # Захват скриншота окна
            img = capture_window_image(hwnd, client_width, client_height)
            # Обработка каждой области
            for team, groups in coord_map.items():
                for group, zones in groups.items():
                    for zone, (left_px, top_px, right_px, bottom_px) in zones.items():
                        region = img.crop((left_px, top_px, right_px, bottom_px))
                        # Определяем путь файла для записи результата
                        out_file = os.path.join(data_dir, team, (group if group != "_global" else ""), 
                                                 "gamer.txt" if zone == "pers" and group != "_global" else f"{zone}.txt")
                        # Если зона относится к изображениям (Q, E, слоты s1-s6)
                        if zone.startswith("s") or zone in ["q", "e"]:
                            zone_type = "s" if zone.startswith("s") else zone
                            match_name = match_image(region, ref_images.get(zone_type, []))
                            if match_name:
                                try:
                                    with open(out_file, "w", encoding="utf-8") as f:
                                        f.write(match_name)
                                except Exception as e:
                                    print(f"Ошибка записи {out_file}: {e}")
                            # Если совпадение не найдено, не изменяем файл
                        else:
                            # OCR для текстовых полей
                            numeric = zone in ["kills", "deaths", "assist", "gold", "gold.global", "lord", "turtle", "tower"]
                            text = ocr_image(region, numeric=numeric)
                            if text:
                                try:
                                    with open(out_file, "w", encoding="utf-8") as f:
                                        f.write(text)
                                except Exception as e:
                                    print(f"Ошибка записи {out_file}: {e}")
                            # Если текст не распознан, файл не трогаем
            # Задержка до следующего цикла ~3 секунды
            elapsed = time.time() - start_time
            if elapsed < 3.0:
                time.sleep(3.0 - elapsed)
    except KeyboardInterrupt:
        print("Мониторинг остановлен пользователем.")
    except Exception as e:
        print(f"Ошибка во время выполнения: {e}")
